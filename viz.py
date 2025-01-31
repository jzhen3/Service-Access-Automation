import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

def show_viz():
    st.header("Explore the Whole Query")
    sheet = openpyxl.load_workbook('employee_service_access_data.xlsx')['Sheet1']
    table = [r for r in sheet.iter_rows(values_only=True)][1:]
    cols = ['Team', 'Job Level', 'Job Role', 'Tool', 'Approval Status']
    df = pd.DataFrame(data=table, columns = cols).reset_index()


    sheet2 = openpyxl.load_workbook('query.xlsx')['Sheet1']
    table2 = [r for r in sheet2.iter_rows(values_only=True)][1:]
    query_cols = ['Team', 'Job Level', 'Job Role', 'Tool']
    df2 = pd.DataFrame(data=table2, columns = query_cols)

    df2 = st.dataframe(df2)

    draw_df = df[['Tool', 'Job Role']]
    app_counts = draw_df.groupby('Job Role')['Tool'].count().reset_index()
    app_counts.columns = ['Job Role', 'App Count']

    st.header("Explore the Usage by Different Teams")
    st.write("Bar graph displaying the number of apps by Teams:")

    fig, ax = plt.subplots()

    ax.bar(app_counts['Job Role'], app_counts['App Count'])
    ax.set_xlabel("Job Role")
    for label in ax.get_xticklabels():
        label.set_rotation(45)
        label.set_ha('right')
    ax.set_ylabel("Counts")
    ax.set_title('Number of Apps by Different Job Roles')
    
    st.pyplot(fig)