import streamlit as st
import pickle
import pandas as pd
import openpyxl

sheet = openpyxl.load_workbook('employee_service_access_data.xlsx')['Sheet1']
table = [r for r in sheet.iter_rows(values_only=True)][1:]
cols = ['Team', 'Job Level', 'Job Role', 'Tool', 'Approval Status']
df = pd.DataFrame(data=table, columns = cols).reset_index()

def load_model():
    with open('catboost.pkl', 'rb') as f:
        loaded_model = pickle.load(f)
    return loaded_model

def show_prediction():
    st.title("Employee Service Access Identification")
    st.write("""### Type the employee's information""")

    teams = ('Analytics', 'Engineering')

    job_levels = ('Intern', 'Associate', 'Senior', 'Manager')

    job_roles = ('Data Scientist',
                'Product Manager',
                'Business Analyst',
                'BI Developer',
                'DevOps Engineer',
                'App Developer',
                'UX Designer',
                'Data Engineer', 
                'QA Developer')

    team = st.selectbox("Team", teams)
    job_level = st.selectbox("Job Level", job_levels)
    job_role = st.selectbox("Job Role", job_roles)

    st.empty()
    st.write("")
    st.empty()

    X_cols = ['Team', 'Job Level', 'Job Role', 'Tool']
    
    app_name = st.selectbox("App Name", tuple(set(df['Tool'])))

    b1 = st.button("Identify the Employee's Service Access Needs")
    if b1:
        X = pd.DataFrame(data=[[team, job_level, job_role, app_name]], columns=X_cols)
        clf =load_model()
        # the classifier prediction result is a 1d array
        label = clf.predict(X)[0]
        status = 'Approved' if label == 1 else 'Denied'
        status_name = "Approval" if label == 1 else "Deinal"
        pred_proba = clf.predict_proba(X)
        if round(pred_proba[0][1], 5) > 0.5:
            prob = round(pred_proba[0][1], 5)
        else:
            prob = round(pred_proba[0][0], 5)
        st.markdown(f"<font size='3'>The service request will be: </font><font size='5'>{status}, {status_name} rate: {prob}</font>", unsafe_allow_html=True)

    st.empty()
    st.empty()
    st.empty()
    st.subheader("Please Select an Application that You'd Like to Query the Information From")

    b2 = st.button("Query the Past Service Acccess Logs")
    sheet2 = openpyxl.load_workbook('query.xlsx')['Sheet1']
    table2 = [r for r in sheet2.iter_rows(values_only=True)][1:]
    df2 = pd.DataFrame(data=table2, columns = X_cols)
    name = df2[(df2['Team'] == team) & (df2['Job Level'] == job_level) & (df2['Job Role'] ==job_role)].Tool
    
    if b2:
        st.markdown(f"<font size='3'>Historic Approved Services within specific group:\n\n</font><font size='5'>{list(name)[0]}</font>", unsafe_allow_html=True)